import copy
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET
from openpyxl import load_workbook as load_excel_workbook

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", NS)


# Trả về tên cột và số dòng từ địa chỉ ô Excel.
def split_ref(ref):
    match = re.fullmatch(r"([A-Z]+)(\d+)", ref)
    return match.group(1), int(match.group(2))


# Đọc workbook OOXML mà không phụ thuộc thư viện Excel bên ngoài.
def load_workbook(source):
    with zipfile.ZipFile(source) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


# Tạo bảng ánh xạ tên worksheet sang file XML tương ứng.
def worksheet_paths(files):
    workbook = ET.fromstring(files["xl/workbook.xml"])
    relations = ET.fromstring(files["xl/_rels/workbook.xml.rels"])
    targets = {item.attrib["Id"]: item.attrib["Target"] for item in relations}
    result = {}
    for sheet in workbook.find(f"{{{NS}}}sheets"):
        relation_id = sheet.attrib[f"{{{REL_NS}}}id"]
        result[sheet.attrib["name"]] = "xl/" + targets[relation_id].lstrip("/")
    return result


# Quản lý shared string để nội dung mới không làm mất định dạng workbook.
class SharedStrings:
    # Đọc danh sách shared string và tạo index tra cứu nhanh.
    def __init__(self, payload):
        self.root = ET.fromstring(payload)
        self.values = [
            "".join(node.text or "" for node in item.iter(f"{{{NS}}}t"))
            for item in self.root.findall(f"{{{NS}}}si")
        ]
        self.index = {value: position for position, value in enumerate(self.values)}

    # Trả về index hiện có hoặc thêm chuỗi mới vào sharedStrings.xml.
    def add(self, value):
        if value in self.index:
            return self.index[value]
        item = ET.SubElement(self.root, f"{{{NS}}}si")
        text = ET.SubElement(item, f"{{{NS}}}t")
        text.text = value
        position = len(self.values)
        self.values.append(value)
        self.index[value] = position
        self.root.set("count", str(position + 1))
        self.root.set("uniqueCount", str(position + 1))
        return position

    # Xuất shared string XML sau khi cập nhật.
    def serialize(self):
        return ET.tostring(self.root, encoding="utf-8", xml_declaration=True)


# Ghi một ô dạng shared string và tái sử dụng style của cột tương ứng.
def set_cell(root, shared, ref, value, style_row=5):
    column, row_number = split_ref(ref)
    sheet_data = root.find(f"{{{NS}}}sheetData")
    row = sheet_data.find(f"{{{NS}}}row[@r='{row_number}']")
    if row is None:
        row = ET.Element(f"{{{NS}}}row", {"r": str(row_number)})
        insert_at = len(sheet_data)
        for index, current in enumerate(sheet_data):
            if int(current.attrib["r"]) > row_number:
                insert_at = index
                break
        sheet_data.insert(insert_at, row)
    cell = row.find(f"{{{NS}}}c[@r='{ref}']")
    if cell is None:
        cell = ET.Element(f"{{{NS}}}c", {"r": ref})
        template = root.find(f".//{{{NS}}}c[@r='{column}{style_row}']")
        if template is not None and "s" in template.attrib:
            cell.set("s", template.attrib["s"])
        row.append(cell)
        row[:] = sorted(row, key=lambda item: split_ref(item.attrib["r"])[0])
    cell.set("t", "s")
    for child in list(cell):
        cell.remove(child)
    value_node = ET.SubElement(cell, f"{{{NS}}}v")
    value_node.text = str(shared.add(value))


# Ghi trọn một dòng mapping gồm chín cột theo cấu trúc workbook hiện tại.
def set_mapping_row(root, shared, row_number, values):
    for column, value in zip("ABCDEFGHI", values):
        set_cell(root, shared, f"{column}{row_number}", value)


# Cập nhật nội dung 5 luồng theo API và logic đang tồn tại trong source code.
def update_flows(files):
    shared = SharedStrings(files["xl/sharedStrings.xml"])
    paths = worksheet_paths(files)
    sheets = {name: ET.fromstring(files[path]) for name, path in paths.items()}

    overview = sheets["Tổng quan"]
    set_cell(overview, shared, "B5", "Xác thực, hồ sơ và thanh toán tài nguyên")
    set_cell(overview, shared, "C5", "12")
    set_cell(overview, shared, "C6", "9")

    flow1 = sheets["Luồng 1"]
    set_cell(flow1, shared, "A1", "Luồng 1 — Xác thực, hồ sơ và thanh toán tài nguyên: Chức năng ↔ API ↔ Controller ↔ Logic ↔ FE gọi ở đâu")
    payment_rows = [
        (11, ["7", "Xem các bundle tài nguyên", "Guest/User", "GET /api/subscription/plans", "backend/src/payments/payments.controller.ts → getPlans()", "PaymentsService.getPlans() trả giá và snapshot quyền lợi cộng thêm: 30 ngày, storage, upload credit, AI credit và số ngày AI không giới hạn.", "frontend/src/api/payments.api.ts → SubscriptionView.tsx", "Đã cập nhật", "Student/Pro là bundle tài nguyên, không còn là trạng thái loại trừ nhau."]),
        (12, ["8", "Xem ví tài nguyên hiện tại", "User", "GET /api/subscription/current", "backend/src/payments/payments.controller.ts → getCurrentSubscription()", "Xác thực Firebase → kiểm tra accessExpiresAt → tính dung lượng, upload và AI còn lại; quyền AI không giới hạn dùng unlimitedAiUntil riêng.", "frontend/src/api/payments.api.ts → SubscriptionView.tsx", "Đã cập nhật", "Nếu quyền trả phí hết hạn, hệ thống khôi phục mức Free."]),
        (13, ["9", "Tạo đơn mua thêm tài nguyên", "User", "POST /api/payments/checkout", "backend/src/payments/payments.controller.ts → createCheckout()", "Kiểm tra bundle/phương thức → chốt giá và quyền lợi vào PaymentOrder → tạo form ký HMAC cho SePay; cho phép mua Student khi Pro còn hiệu lực.", "frontend/src/api/payments.api.ts → SubscriptionView.tsx", "Đã cập nhật", "Payment lưu durationDays, storageMb, uploadCredits, aiCredits và unlimitedAiDays."]),
        (14, ["10", "Nhận kết quả thanh toán SePay", "SePay", "POST /api/payments/sepay/ipn", "backend/src/payments/payments.controller.ts → processIpn()/processWebhook()", "Xác thực API key và đối soát invoice/số tiền → đánh dấu PAID idempotent → cộng 30 ngày từ max(now, expiresAt) → cộng storage/upload/AI → ghi EntitlementTransaction và audit log trong cùng transaction.", "SePay gọi webhook; FE không gọi trực tiếp", "Đã cập nhật", "Một payment chỉ được cộng tài nguyên đúng một lần."]),
        (15, ["11", "Theo dõi trạng thái thanh toán", "User", "GET /api/payments/:invoiceNumber", "backend/src/payments/payments.controller.ts → getPayment()", "Đọc PaymentOrder; nếu còn PENDING thì đối soát trạng thái CAPTURED với SePay và áp dụng tài nguyên như webhook dự phòng.", "frontend/src/api/payments.api.ts → SubscriptionView.tsx", "Đã có", "Redirect thành công không tự cấp tài nguyên nếu chưa xác nhận từ gateway."]),
        (16, ["12", "Xem lịch sử hoặc hủy phiên thanh toán", "User", "GET /api/payments/history; POST /api/payments/:invoiceNumber/status", "backend/src/payments/payments.controller.ts", "Trả lịch sử giao dịch; tự đánh dấu order quá hạn; chỉ cho cập nhật FAILED/CANCELLED khi payment chưa PAID.", "frontend/src/api/payments.api.ts → SubscriptionView.tsx", "Đã có", "Refund đảo đúng entitlement của payment, không mặc định xóa toàn bộ quyền lợi khác."]),
    ]
    for row_number, values in payment_rows:
        set_mapping_row(flow1, shared, row_number, values)

    flow2 = sheets["Luồng 2"]
    upload_rows = [
        (5, ["1", "FE gửi file và metadata", "User", "POST /api/documents (multipart/form-data)", "backend/src/documents/documents.controller.ts → upload()", "FirebaseAuthGuard → ParseFilePipe kiểm tra type/size → validateUpload() → DocumentsService.upload() → startExtraction(). Đây là API chính mà UploadDocumentView gọi.", "frontend/src/api/documents.api.ts → UploadDocumentView.tsx", "Đã có", "FE không gửi file trực tiếp đến Cloudflare."]),
        (6, ["2", "Backend upload file lên Cloudflare R2", "System", "Không có API FE riêng; được kích hoạt bởi POST /api/documents", "backend/src/documents/documents.service.ts → upload(); backend/src/storage/storage.service.ts → uploadObject()", "DocumentsService gọi StorageService.uploadObject(ownerId, file); StorageService tạo object key và dùng S3-compatible PutObject tới bucket R2 private; sau đó lưu storagePath/fileUrl vào documents.", "Không gọi riêng từ FE", "Đã có", "POST /api/storage/upload-url tồn tại cho presigned flow khác nhưng UploadDocumentView hiện không sử dụng."]),
        (7, ["3", "Tạo metadata và bản ghi content chờ xử lý", "System", "Nằm trong POST /api/documents", "backend/src/documents/documents.service.ts → upload()", "Prisma transaction create Document cùng DocumentContent có extractionStatus=PENDING, sourceType và progress=0; tài liệu PUBLIC có moderationStatus=PENDING.", "Không gọi riêng từ FE", "Đã có", "Đây là bước tạo document_contents ban đầu."]),
        (8, ["4", "Xếp job content extraction", "System/User", "Tự động sau upload; POST /api/documents/:id/extract để chạy lại", "backend/src/document-content/document-content.controller.ts → extract(); ContentExtractionService.startExtraction()", "Kiểm tra quyền owner và trạng thái job → tạo jobId/đặt PROCESSING → đưa công việc vào ExtractionQueueService.", "Upload tự kích hoạt; LibraryView có thể gọi retry", "Đã có", "API extract không nhận lại file; worker đọc object từ R2."]),
        (9, ["5", "Trích xuất và ghi content vào database", "System worker", "Không có API ghi DB công khai", "backend/src/content-extraction/content-extraction.service.ts → processExtraction()", "Worker tải file từ R2 → chọn PDF/DOCX/PPTX/XLSX extractor → OCR fallback nếu cần → update document_contents.extractedText/quality/progress/status → replaceVectorChunks() ghi document_chunks và embedding.", "Không gọi trực tiếp từ FE", "Đã có", "Toàn bộ ghi DB nằm trong service/Prisma, không mở endpoint cho client nhập extracted content."]),
        (10, ["6", "Đọc nội dung đã trích xuất", "User", "GET /api/documents/:id/content", "backend/src/document-content/document-content.controller.ts → getContent()", "Xác thực quyền truy cập tài liệu → trả extractedText, source type, chất lượng và metadata xử lý đã lưu trong document_contents.", "frontend gọi qua documents API khi cần", "Đã có", "Đây là API đọc content, không phải API ghi extraction."]),
        (11, ["7", "Theo dõi trạng thái extraction", "User", "GET /api/documents/:id/extraction-status", "backend/src/document-content/document-content.controller.ts → getStatus()", "Trả PENDING/PROCESSING/COMPLETED/FAILED, progress, jobId và lỗi để FE polling hoặc hiển thị retry.", "frontend/src/api/documents.api.ts → LibraryView.tsx", "Đã có", "Phân biệt với GET /api/documents dùng tải danh sách thư viện."]),
        (12, ["8", "Quét rủi ro để hỗ trợ admin", "System", "Pipeline nội bộ", "backend/src/content-extraction/moderation-scanner.service.ts", "Quét title/file/content → ghi moderationFlag, priority, matchedKeywords và matchedContexts; bộ quét không được tự APPROVED tài liệu PUBLIC.", "Không gọi trực tiếp từ FE", "Đã cập nhật", "Mọi tài liệu công khai vẫn chờ admin duyệt ở Luồng 5."]),
        (13, ["9", "Preview hoặc download tài liệu", "User", "GET /api/documents/:id/preview; GET /api/documents/:id/download", "backend/src/documents/documents.controller.ts", "Kiểm tra quyền → StorageService tạo presigned URL ngắn hạn tới R2; download đồng thời tăng counter và ghi DownloadLog.", "frontend/src/api/documents.api.ts → LibraryView.tsx", "Đã có", "Office file có thể dùng Microsoft Office Viewer fallback."]),
    ]
    for row_number, values in upload_rows:
        set_mapping_row(flow2, shared, row_number, values)

    flow5 = sheets["Luồng 5"]
    set_mapping_row(flow5, shared, 5, ["1", "User gửi yêu cầu công khai", "User", "PUT /api/documents/:id/visibility", "backend/src/documents/documents.controller.ts → updateVisibility()", "Kiểm tra owner → chuyển visibility=PUBLIC và bắt buộc moderationStatus=PENDING; nội dung scanner chỉ cung cấp flag/priority, không tự duyệt.", "frontend/src/api/documents.api.ts → LibraryView.tsx", "Đã cập nhật", "Nút thư viện hiển thị Chờ kiểm duyệt; user có thể hủy để chuyển PRIVATE."])
    set_mapping_row(flow5, shared, 6, ["2", "Admin xem hàng chờ kiểm duyệt", "Admin", "GET /api/admin/documents", "backend/src/admin/admin-documents.controller.ts → findAll()", "FirebaseAuthGuard + RolesGuard(ADMIN) → chỉ lấy visibility=PUBLIC → lọc moderationStatus/flag/keyword → ưu tiên moderationPriority và submittedAt.", "frontend/src/api/admin.api.ts → AdminDocumentsView.tsx", "Đã có", "Private documents không xuất hiện trong moderation console."])
    set_mapping_row(flow5, shared, 7, ["3", "Admin preview tài liệu", "Admin", "GET /api/admin/documents/:id/preview", "backend/src/admin/admin-documents.controller.ts → preview()", "Kiểm tra tài liệu PUBLIC → tạo presigned preview URL từ StorageService để admin đọc nội dung trước khi quyết định.", "frontend/src/api/admin.api.ts → AdminDocumentsView.tsx", "Đã có", "Không thay đổi trạng thái khi preview."])
    set_mapping_row(flow5, shared, 8, ["4", "Admin duyệt tài liệu", "Admin", "PUT /api/admin/documents/:id/approve", "backend/src/admin/admin-documents.controller.ts → approve()", "Kiểm tra PUBLIC → set moderationStatus=APPROVED, status=ACTIVE, clear rejectionReason, lưu reviewedAt/reviewedBy → audit log và notification DOCUMENT_APPROVED.", "frontend/src/api/admin.api.ts → AdminDocumentsView.tsx", "Đã có", "Chỉ sau bước này tài liệu mới xuất hiện ở Community."])
    set_mapping_row(flow5, shared, 9, ["5", "Admin từ chối tài liệu", "Admin", "PUT /api/admin/documents/:id/reject", "backend/src/admin/admin-documents.controller.ts → reject()", "Bắt buộc reason → set REJECTED/HIDDEN, lưu reviewer và lý do → audit log → notification DOCUMENT_REJECTED cho owner.", "frontend/src/api/admin.api.ts → AdminDocumentsView.tsx", "Đã có", "Owner nhìn thấy lý do và có thể chỉnh sửa/gửi duyệt lại."])
    set_mapping_row(flow5, shared, 10, ["6", "Community chỉ đọc tài liệu đã duyệt", "Guest/User", "GET /api/community/documents", "backend/src/community/community.service.ts → findDocuments()", "Query bắt buộc visibility=PUBLIC, moderationStatus=APPROVED và status=ACTIVE; PENDING/REJECTED không thể bị lộ qua API cộng đồng.", "frontend/src/api/community.api.ts → CommunityView.tsx", "Đã có", "Đây là lớp bảo vệ cuối ngoài giao diện admin."])
    set_mapping_row(flow5, shared, 11, ["7", "Theo dõi audit và thông báo", "System/Admin", "GET /api/audit-logs; GET /api/notifications", "backend/src/audit-log; backend/src/notifications", "Lưu actor/action/metadata cho quyết định kiểm duyệt và gửi trạng thái APPROVED/REJECTED/PENDING tới người sở hữu tài liệu.", "AdminDashboardView.tsx; notifications.api.ts", "Đã có", "Đảm bảo truy vết đầy đủ quyết định của admin."])

    for name, root in sheets.items():
        files[paths[name]] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    files["xl/sharedStrings.xml"] = shared.serialize()


# Ghi workbook mới và giữ nguyên toàn bộ entry không bị chỉnh sửa.
def save_workbook(files, destination):
    with zipfile.ZipFile(destination, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, payload in files.items():
            archive.writestr(name, payload)
    normalized = destination.with_suffix(".normalized.xlsx")
    workbook = load_excel_workbook(destination)
    workbook.save(normalized)
    normalized.replace(destination)


# Điều phối việc đọc, cập nhật và ghi workbook từ tham số dòng lệnh.
def main():
    source = Path(sys.argv[1])
    destination = Path(sys.argv[2])
    files = load_workbook(source)
    update_flows(files)
    save_workbook(files, destination)
    print(destination)


if __name__ == "__main__":
    main()
